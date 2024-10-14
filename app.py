from io import BytesIO
from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from constants import (
    CONFIRMED, TOTAL, UNMARRIED, MEDICAL_STATE, SICK, HOSPITALIZED, UNKNOWN, SIGNS_OBSERVED, ICD_9_DIAGNOSIS,
    VACCINATED_FOR_FLU, VACCINATED_FOR_PNEUMOCOCCUS, SUSPECTED, POSSIBLE
)

from utils import create_data_row
from openpyxl.styles import Border, Side

app = Flask(__name__)


def identify_csv_file(file):
    """Identify the CSV file based on its column names."""
    df = pd.read_csv(file)

    # Check for unique columns to identify the file
    if 'totsheruarmundshem' in df.columns:
        return 'df1', df
    elif 'icd9mundshempo' in df.columns:
        return 'df2', df
    elif 'totalmundshem_pacient' in df.columns:
        return 'df3', df
    else:
        raise ValueError("Unrecognized CSV structure")


def map_data_from_all_csvs(df1, df2, df3):
    """Function to map data from the three CSVs into the correct structure."""
    mapped_data = {
        CONFIRMED: {
            TOTAL: df3['totalkonfirmuar_pacient'].sum(),
            'GJINIA': {
                'MESHKUJ': df3['konfirmuarpacientm'].sum(),
                'FEMRA': df3['konfirmuarpacientf'].sum(),
            },
            'MOSHA': {
                '0-1': {'MESHKUJ': df3['mk0_1'].sum(), 'FEMRA': df3['fk0_1'].sum()},
                '1-18': {'MESHKUJ': df3['mk1_18'].sum(), 'FEMRA': df3['fk1_18'].sum()},
                '19-25': {'MESHKUJ': df3['mk19_25'].sum(), 'FEMRA': df3['fk19_25'].sum()},
                '26-34': {'MESHKUJ': df3['mk26_34'].sum(), 'FEMRA': df3['fk26_34'].sum()},
                '35-70': {'MESHKUJ': df3['mk35_70'].sum(), 'FEMRA': df3['fk35_70'].sum()},
                '>70': {'MESHKUJ': df3['mkmbi70'].sum(), 'FEMRA': df3['fkmbi70'].sum()},
            },
            'STATUSI': {
                'MARTUAR': df3['totalkonfimuar_pacientmartuar'].sum(),
                UNMARRIED: df3['totalkonfimuar_pacientpamartuar'].sum(),
            },
            MEDICAL_STATE: {
                'SHERUAR': df1['totsheruarkonfirmuar'].sum(),
                SICK: df1['totpasheruarkonfirmuar'].sum(),
                'VDEKUR': df1['totvdekurkonfirmuar'].sum(),
                'PANJOHUR': df1['totpanjohurkonfirmuar'].sum(),
            },
            HOSPITALIZED: {
                'PO': df1['totshtrimpokonfirmuar'].sum(),
                'JO': df1['totshtrimjokonfirmuar'].sum(),
                UNKNOWN: df1['totshtrimpanjohurkonfirmuar'].sum(),
            },
            'SIMPTOMA': {
                'PO': df1['totsimptomapokonfirmuar'].sum(),
                'JO': df1['totsimptomajokonfirmuar'].sum(),
                UNKNOWN: df1['totsimptomapanjohurkonfirmuar'].sum(),
            },
            SIGNS_OBSERVED: {
                'PO': df1['totshenjapokonfirmuar'].sum(),
                'JO': df1['totshenjajokonfirmuar'].sum(),
                UNKNOWN: df1['totshenjapanjohurkonfirmuar'].sum(),
            },
            ICD_9_DIAGNOSIS: {
                'PO': df2['icd9konfirmuarpo'].sum(),
                'JO': df2['icd9konfirmuarjo'].sum(),
            },
            VACCINATED_FOR_FLU: {
                'PO': df1['totvaksinuargrippokonfirmuar'].sum(),
                'JO': df1['totvaksinuargripjokonfirmuar'].sum(),
                UNKNOWN: df1['totvaksinuargrippanjohurkonfirmuar'].sum(),
            },
            VACCINATED_FOR_PNEUMOCOCCUS: {
                'PO': df1['totvaksinuarpneumopokonfirmuar'].sum(),
                'JO': df1['totvaksinuarpneumojokonfirmuar'].sum(),
                UNKNOWN: df1['totvaksinuarpneumopanjohurkonfirmuar'].sum(),
            },
        },
        SUSPECTED: {
            TOTAL: df3['totaldyshuar_pacient'].sum(),
            'GJINIA': {
                'MESHKUJ': df3['dyshuarpacientm'].sum(),
                'FEMRA': df3['dyshuarpacientf'].sum(),
            },
            'MOSHA': {
                '0-1': {'MESHKUJ': df3['md0_1'].sum(), 'FEMRA': df3['fd0_1'].sum()},
                '1-18': {'MESHKUJ': df3['md1_18'].sum(), 'FEMRA': df3['fd1_18'].sum()},
                '19-25': {'MESHKUJ': df3['md19_25'].sum(), 'FEMRA': df3['fd19_25'].sum()},
                '26-34': {'MESHKUJ': df3['md26_34'].sum(), 'FEMRA': df3['fd26_34'].sum()},
                '35-70': {'MESHKUJ': df3['md35_70'].sum(), 'FEMRA': df3['fd35_70'].sum()},
                '>70': {'MESHKUJ': df3['mdmbi70'].sum(), 'FEMRA': df3['fdmbi70'].sum()},
            },
            'STATUSI': {
                'MARTUAR': df3['totaldyshuar_pacientmartuar'].sum(),
                UNMARRIED: df3['totaldyshuar_pacientpamartuar'].sum(),
            },
            MEDICAL_STATE: {
                'SHERUAR': df1['totsheruardyshuar'].sum(),
                SICK: df1['totpasheruardyshuar'].sum(),
                'VDEKUR': df1['totvdekurdyshuar'].sum(),
                'PANJOHUR': df1['totpanjohurdyshuar'].sum(),
            },
            HOSPITALIZED: {
                'PO': df1['totshtrimpodyshuar'].sum(),
                'JO': df1['totshtrimjodyshuar'].sum(),
                UNKNOWN: df1['totshtrimpanjohurdyshuar'].sum(),
            },
            'SIMPTOMA': {
                'PO': df1['totsimptomapodyshuar'].sum(),
                'JO': df1['totsimptomajodyshuar'].sum(),
                UNKNOWN: df1['totsimptomapanjohurdyshuar'].sum(),
            },
            SIGNS_OBSERVED: {
                'PO': df1['totshenjapodyshuar'].sum(),
                'JO': df1['totshenjajodyshuar'].sum(),
                UNKNOWN: df1['totshenjapanjohurdyshuar'].sum(),
            },
            ICD_9_DIAGNOSIS: {
                'PO': df2['icd9dyshuarpo'].sum(),
                'JO': df2['icd9dyshuarjo'].sum(),
                UNKNOWN: 0,  # Placeholder for unknown values
            },
            VACCINATED_FOR_FLU: {
                'PO': df1['totvaksinuargrippodyshuar'].sum(),
                'JO': df1['totvaksinuargripjodyshuar'].sum(),
                UNKNOWN: df1['totvaksinuargrippanjohurdyshuar'].sum(),
            },
            VACCINATED_FOR_PNEUMOCOCCUS: {
                'PO': df1['totvaksinuarpneumopodyshuar'].sum(),
                'JO': df1['totvaksinuarpneumojodyshuar'].sum(),
                UNKNOWN: df1['totvaksinuarpneumopanjohurdyshuar'].sum(),
            },
        },
        POSSIBLE: {
            TOTAL: df3['totalmundshem_pacient'].sum(),
            'GJINIA': {
                'MESHKUJ': df3['mundshempacientm'].sum(),
                'FEMRA': df3['mundshempacientf'].sum(),
            },
            'MOSHA': {
                '0-1': {'MESHKUJ': df3['mm0_1'].sum(), 'FEMRA': df3['fm0_1'].sum()},
                '1-18': {'MESHKUJ': df3['mm1_18'].sum(), 'FEMRA': df3['fm1_18'].sum()},
                '19-25': {'MESHKUJ': df3['mm19_25'].sum(), 'FEMRA': df3['fm19_25'].sum()},
                '26-34': {'MESHKUJ': df3['mm26_34'].sum(), 'FEMRA': df3['fm26_34'].sum()},
                '35-70': {'MESHKUJ': df3['mm35_70'].sum(), 'FEMRA': df3['fm35_70'].sum()},
                '>70': {'MESHKUJ': df3['mmmbi70'].sum(), 'FEMRA': df3['fmmbi70'].sum()},
            },
            'STATUSI': {
                'MARTUAR': df3['totalmundshem_pacientmartuar'].sum(),
                UNMARRIED: df3['totalmundshem_pacientpamartuar'].sum(),
            },
            MEDICAL_STATE: {
                'SHERUAR': df1['totsheruarmundshem'].sum(),
                SICK: df1['totpasheruarmundshem'].sum(),
                'VDEKUR': df1['totvdekurmundshem'].sum(),
                'PANJOHUR': df1['totpanjohurmundshem'].sum(),
            },
            HOSPITALIZED: {
                'PO': df1['totshtrimpomundshem'].sum(),
                'JO': df1['totshtrimjomundshem'].sum(),
                UNKNOWN: df1['totshtrimpanjohurmundshem'].sum(),
            },
            'SIMPTOMA': {
                'PO': df1['totsimptomapomundshem'].sum(),
                'JO': df1['totsimptomajomundshem'].sum(),
                UNKNOWN: df1['totsimptomapanjohurmundshem'].sum(),
            },
            SIGNS_OBSERVED: {
                'PO': df1['totshenjapomundshem'].sum(),
                'JO': df1['totshenjajomundshem'].sum(),
                UNKNOWN: df1['totshenjapanjohurmundshem'].sum(),
            },
            ICD_9_DIAGNOSIS: {
                'PO': df2['icd9mundshempo'].sum(),
                'JO': df2['icd9mundshemjo'].sum(),
                UNKNOWN: 0,  # Placeholder for unknown values
            },
            VACCINATED_FOR_FLU: {
                'PO': df1['totvaksinuargrippomundshem'].sum(),
                'JO': df1['totvaksinuargripjomundshem'].sum(),
                UNKNOWN: df1['totvaksinuargrippanjohurmundshem'].sum(),
            },
            VACCINATED_FOR_PNEUMOCOCCUS: {
                'PO': df1['totvaksinuarpneumopomundshem'].sum(),
                'JO': df1['totvaksinuarpneumojomundshem'].sum(),
                UNKNOWN: df1['totvaksinuarpneumopanjohurmundshem'].sum(),
            },
        },
    }

    return mapped_data


def write_mapped_data_to_excel(mapped_data, output_path):
    # Create a workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Set column widths for a cleaner look using column letters
    column_widths = [30, 10, 15, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 20, 15, 15, 15, 10, 10, 10, 10, 10,
                     10, 10, 10, 10, 10, 10, 20, 20, 10, 15, 15, 10, 10]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)  # Convert index to column letter (A, B, C, ..., Z, AA, AB, ...)
        ws.column_dimensions[col_letter].width = width

    # Manually setting up the header
    header = [
        # First row headers
        ["RASTET COVID-19", "NR TOTAL", "GJINIA", "", "MOSHA", "", "", "", "", "", "", "", "", "", "", "",
         "STATUSI", "", "GJENDJA SHENDETESORE", "", "", "", "SHTRIM NE SPITAL", "", "", "SIMPTOMA", "", "",
         "SHENJA TE OBSERVUARA", "", "", "Semundje bashkeshoqeruese sipas diagnozave icd9", "", "VAKSINUAR PER GRIPIN",
         "", "", "VAKSINUAR PER PNEUMOKOK", "", ""
         ],
        # Second row headers
        ["", "", "", "", "0-1 years", "", "1-18 years", "", "19-25 years", "", "26-34 years", "", "35-70 years",
         "", ">70 years", ] + [""] * 23,
        # Third row headers
        ["", "", ] + ["MESHKUJ", "FEMRA"] * 7
        + ["MARTUAR", "JO MARTUAR", "SHERUAR", "I PASHERUAR", "VDEKUR", "PANJOHUR"] + ["PO", "JO", UNKNOWN, ] * 3
        + ["PO", "JO"]
        + ["PO", "JO", "PANJOHUR"] * 2
    ]

    # Define a border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Adding the header
    for row in header:
        ws.append(row)

    # Apply styling (bold headers, center alignment, etc.)
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=len(header[0])):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Merge cells to match the structure
    ws.merge_cells('A1:A3')  # COVID-19 Cases
    ws.merge_cells('B1:B3')  # TOTAL
    ws.merge_cells('C1:D2')  # Gender
    ws.merge_cells('E1:P1')  # Age
    ws.merge_cells('Q1:R1')  # Status
    ws.merge_cells('S1:V1')  # Medical state
    ws.merge_cells('W1:Y1')  # Hospitalized
    ws.merge_cells('Z1:AB1')  # Symptoms
    ws.merge_cells('AC1:AE1')  # Observed signs
    ws.merge_cells('AF1:AG1')  # icd9 diseases
    ws.merge_cells('AH1:AJ1')  # Vacc for flu
    ws.merge_cells('AK1:AM1')  # Vacc for pneumococcus
    ws.merge_cells('E2:F2')  # 0-1 years
    ws.merge_cells('G2:H2')  # 0-1 years
    ws.merge_cells('I2:J2')  # 0-1 years
    ws.merge_cells('K2:L2')  # 0-1 years
    ws.merge_cells('M2:N2')  # 0-1 years
    ws.merge_cells('O2:P2')  # 0-1 years

    # Create the data rows for CONFIRMED, SUSPECTED, and POSSIBLE cases
    data_rows = [
        create_data_row(CONFIRMED, mapped_data),
        create_data_row(SUSPECTED, mapped_data),
        create_data_row(POSSIBLE, mapped_data)
    ]

    # Add the data rows to the worksheet
    for row in data_rows:
        ws.append(row)

    # Apply styling (bold headers, center alignment, etc.)
    for row in ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=len(header[0])):
        for cell in row:
            cell.border = thin_border

    # Save the workbook
    wb.save(output_path)
    print("Excel structure created and saved!")


def generate_html_table(mapped_data):
    """Function to generate an HTML table from the mapped data, with merged cells like in Excel."""

    # Define the header structure, with merged cells
    header_rows = [
        # Row 1: Top-level headers with merged cells
        [
            ('RASTET COVID-19', 1, 3),  # (content, colspan, rowspan)
            ('NR TOTAL', 1, 3),
            ('GJINIA', 2, 2),
            ('MOSHA', 12, 1),
            ('STATUSI', 2, 1),
            ('GJENDJA SHENDETESORE', 4, 1),
            ('SHTRIM NE SPITAL', 3, 1),
            ('SIMPTOMA', 3, 1),
            ('SHENJA TE OBSERVUARA', 3, 1),
            ('Semundje bashkeshoqeruese sipas diagnozave icd9', 2, 1),
            ('VAKSINUAR PER GRIPIN', 3, 1),
            ('VAKSINUAR PER PNEUMOKOK', 3, 1)
        ],
        # Row 2: Sub-headers
        [
            ('0-1 years', 2, 1), ('1-18 years', 2, 1), ('19-25 years', 2, 1),
            ('26-34 years', 2, 1), ('35-70 years', 2, 1), ('>70 years', 2, 1),
            *[('', 1, 1)] * 23,  # Preceding * to unpack and flatten
        ],
        # Row 3 subheaders
        [
            *[("MESHKUJ", 1, 1), ("FEMRA", 1, 1)] * 7,  # The preceding * unpacks and flattens
            ("MARTUAR", 1, 1), ("JO MARTUAR", 1, 1),
            ("SHERUAR", 1, 1), ("I PASHERUAR", 1, 1), ("VDEKUR", 1, 1), ("PANJOHUR", 1, 1),
            *[("PO", 1, 1), ("JO", 1, 1), (UNKNOWN, 1, 1)] * 3,
            ("PO", 1, 1), ("JO", 1, 1),
            *[("PO", 1, 1), ("JO", 1, 1), (UNKNOWN, 1, 1)] * 2,
        ]
    ]

    # Start HTML table
    html_content = "<table border='1'>"

    # Render the header rows with merged cells using colspan and rowspan
    for header_row in header_rows:
        html_content += "<tr>"
        for cell, colspan, rowspan in header_row:
            if colspan > 1 or rowspan > 1:
                html_content += f"<th colspan='{colspan}' rowspan='{rowspan}'>{cell}</th>"
            else:
                html_content += f"<th>{cell}</th>"
        html_content += "</tr>"

    # Create the data rows for CONFIRMED, SUSPECTED, and POSSIBLE cases
    data_rows = [
        create_data_row(CONFIRMED, mapped_data),
        create_data_row(SUSPECTED, mapped_data),
        create_data_row(POSSIBLE, mapped_data)
    ]

    # Render the data rows
    for row in data_rows:
        html_content += "<tr>"
        for cell in row:
            html_content += f"<td>{cell}</td>"
        html_content += "</tr>"

    html_content += "</table>"

    return html_content


@app.route('/')
def show_main():
    return render_template('upload.html')


@app.route('/upload', methods=['POST'])
def upload():
    """Route to handle file upload, Excel generation, and HTML display."""
    if 'first_file' not in request.files or 'second_file' not in request.files or 'third_file' not in request.files:
        return "Upload all three files"

    first_file = request.files['first_file']
    second_file = request.files['second_file']
    third_file = request.files['third_file']

    # Identifying the CSV files based on their content
    file_mapping = {}

    try:
        csv_file = identify_csv_file(first_file)
        file_mapping[csv_file[0]] = csv_file[1]
        csv_file2 = identify_csv_file(second_file)
        file_mapping[csv_file2[0]] = csv_file2[1]
        csv_file3 = identify_csv_file(third_file)
        file_mapping[csv_file3[0]] = csv_file3[1]
    except ValueError as e:
        return str(e)

    # Extract the identified dataframes
    df1 = file_mapping.get('df1')
    df2 = file_mapping.get('df2')
    df3 = file_mapping.get('df3')

    if df1 is None or df2 is None or df3 is None:
        return "Ensure you have uploaded all 3 different csv files."

    # Proceed with processing the files
    mapped_data = map_data_from_all_csvs(df1, df2, df3)

    # Determine action (download or show content)
    action = request.form.get('action')

    if action == 'download':
        output = BytesIO()
        write_mapped_data_to_excel(mapped_data, output)
        output.seek(0)
        return send_file(output, download_name="processed_file.xlsx", as_attachment=True)

    elif action == 'show':
        html_table = generate_html_table(mapped_data)
        return html_table

    return "Invalid file format. Please upload valid .csv files."


if __name__ == '__main__':
    app.run()
